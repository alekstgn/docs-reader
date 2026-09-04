"""Таблица замечаний для проверяющего — отдельный файл на каждый отчёт."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .issues import Issue
from .locate import build_doc_map, locate_issues

HEADERS = ["№", "Страница", "Где в отчёте", "В чём проблема"]
COL_WIDTHS = [6, 10, 38, 64]
TABLE_HEADER_ROW = 4

FILL_TITLE = PatternFill("solid", fgColor="0B2C4A")
FILL_META = PatternFill("solid", fgColor="E8EEF4")
FILL_HEAD = PatternFill("solid", fgColor="1F4E79")
FONT_TITLE = Font(name="Calibri", size=14, color="FFFFFF", bold=True)
FONT_META = Font(name="Calibri", size=11, color="1A1A1A")
FONT_HEAD = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
FONT_BODY = Font(name="Calibri", size=11, color="1A1A1A")
THIN = Border(
    left=Side(style="thin", color="B0B8C1"),
    right=Side(style="thin", color="B0B8C1"),
    top=Side(style="thin", color="B0B8C1"),
    bottom=Side(style="thin", color="B0B8C1"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _page_sort_key(page: str) -> tuple[int, str]:
    if not page:
        return (1_000_000, "")
    match = re.search(r"\d+", page)
    if match:
        return (int(match.group()), page)
    return (999_999, page)


def _page_cell_value(page: str) -> int | str:
    if not page:
        return ""
    match = re.search(r"\d+", page)
    if match:
        return int(match.group())
    return page


def sort_located_by_page(
    located: list[tuple[Issue, int | None, str]],
) -> list[tuple[Issue, int | None, str]]:
    return sorted(
        located,
        key=lambda item: (_page_sort_key(item[2]), item[0].location, item[0].description),
    )


def format_event_header(event: dict[str, str | bool] | None) -> str:
    """В шапке — мероприятие по договору (эталон сверки), не из отчёта."""
    event = event or {}
    required = str(event.get("required") or "").strip()
    claimed = str(event.get("claimed") or "").strip()
    return required or claimed


def _write_header_block(
    ws,
    *,
    event_summary: str,
    issue_count: int,
) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title = ws.cell(row=1, column=1, value="Список замечаний по сверке отчёта с договором")
    title.fill = FILL_TITLE
    title.font = FONT_TITLE
    title.alignment = CENTER
    title.border = THIN
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    event_cell = ws.cell(row=2, column=1, value=f"Мероприятие: {event_summary}")
    event_cell.fill = FILL_META
    event_cell.font = FONT_META
    event_cell.alignment = WRAP
    event_cell.border = THIN
    ws.row_dimensions[2].height = 30

    formed = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(HEADERS))
    meta = ws.cell(
        row=3,
        column=1,
        value=f"Сформировано: {formed}. Всего замечаний: {issue_count}.",
    )
    meta.fill = FILL_META
    meta.font = FONT_META
    meta.alignment = WRAP
    meta.border = THIN


def _write_table(ws, located: list[tuple[Issue, int | None, str]], *, start_row: int) -> None:
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = FILL_HEAD
        cell.font = FONT_HEAD
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[col - 1]

    ws.freeze_panes = f"A{start_row + 1}"
    last_row = start_row + max(1, len(located))
    ws.auto_filter.ref = f"A{start_row}:D{last_row}"

    for row_num, (issue, _para, page) in enumerate(located, 1):
        row_idx = start_row + row_num
        values = [
            row_num,
            _page_cell_value(page),
            issue.location[:120],
            issue.description,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = FONT_BODY
            cell.alignment = WRAP if col != 2 else Alignment(horizontal="center", vertical="top")
            cell.border = THIN


def export_pair_problems(
    out_path: Path,
    *,
    event_summary: str,
    located: list[tuple[Issue, int | None, str]],
) -> Path:
    ordered = sort_located_by_page(located)
    wb = Workbook()
    ws = wb.active
    ws.title = "Замечания"

    _write_header_block(
        ws,
        event_summary=event_summary,
        issue_count=len(ordered),
    )
    _write_table(ws, ordered, start_row=TABLE_HEADER_ROW)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def locate_report_issues(
    report_path: Path,
    issues: list[Issue],
) -> list[tuple[Issue, int | None, str]]:
    doc_map = build_doc_map(report_path)
    return locate_issues(doc_map, issues)
